"""
importer.py
Fase 4 — file import, in two layers (design: docs/arcos/outbound.md).

  Camada A — batch_ingest(): FORMAT-AGNOSTIC. Normalized rows in → resolve customer_id
             (Identity Resolver) + validate + db_add_entry (upsert by dedup_key) + a
             per-row report out. Never sees a column. Exposed publicly as
             POST /v1/mailings/{id}/entries/batch — reusable by any future format/source.

  Camada B — parse_file(): the file adapter (anti-corruption). Reads the mailing's
             column_map to turn CSV/xlsx columns into Camada-A rows. The only place that
             touches openpyxl / encoding / headers. Delegates ingestion to Camada A.

`entry.metadata` stays OPAQUE at runtime — column_map is a PARSING artifact only.
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Any

import asyncpg

from .db import db_add_entry

logger = logging.getLogger("plughub.mailing.importer")


class ImportParseError(Exception):
    """Unparseable/unsupported file (bad format, missing header) → HTTP 400."""


class RowCapExceeded(Exception):
    """More data rows than the configured cap → HTTP 413."""

    def __init__(self, rows: int, cap: int) -> None:
        super().__init__(f"file has {rows} rows, exceeds cap {cap}")
        self.rows = rows
        self.cap = cap


# ── Camada A — batch ingest (format-agnostic) ─────────────────────────────────

async def batch_ingest(
    pool: asyncpg.Pool,
    tenant_id: str,
    mailing: dict,
    rows: list[dict],
    identity: Any = None,
    resolve: bool = True,
    source: str | None = None,
) -> dict:
    """Ingest normalized rows into a mailing. Per row:
      1. customer_id: use the native id if given; else, when resolve=true and anchors
         are present, resolve via the Identity Resolver (miss → stored raw, counted
         unresolved). A resolve failure never aborts the batch (degrades loud → null).
      2. validate: a row with NO contact address AND NO customer_id is unreachable →
         rejected (never inserted). This is the only rejection reason at this layer.
      3. db_add_entry (upsert by dedup_key derived from the mailing's dedup_policy).
    Returns {total, added, deduped, resolved, unresolved, rejected:[{index, reason}]}.
    """
    added = deduped = resolved = unresolved = 0
    rejected: list[dict] = []

    for i, row in enumerate(rows):
        customer_id = row.get("customer_id")
        anchors     = row.get("anchors") or []
        contacts    = row.get("contacts") or {}
        metadata    = row.get("metadata") or {}

        # 1) resolve (only when there is no native id AND anchors exist AND enabled)
        if not customer_id and resolve and anchors and identity is not None:
            cid = await identity.resolve(tenant_id, anchors)
            if cid:
                customer_id = cid
                resolved += 1
            else:
                unresolved += 1

        # 2) validate reachability
        if not contacts and not customer_id:
            rejected.append({"index": i, "reason": "no reachable contact nor customer_id"})
            continue

        # 3) upsert
        res = await db_add_entry(pool, tenant_id, mailing, {
            "customer_id": customer_id,
            "contacts":    contacts,
            "metadata":    metadata,
            "dedup_key":   row.get("dedup_key"),
            "source":      source,
        })
        if res["deduped"]:
            deduped += 1
        else:
            added += 1

    return {
        "total":      len(rows),
        "added":      added,
        "deduped":    deduped,
        "resolved":   resolved,
        "unresolved": unresolved,
        "rejected":   rejected,
    }


# ── Camada B — file parsing (column_map → rows) ───────────────────────────────

def _decode_csv(content: bytes) -> str:
    """Decode CSV bytes. Try UTF-8 (with BOM), fall back to latin-1 (never raises)."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def _read_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = _decode_csv(content)
    # Sniff the delimiter (',' or ';' — pt-BR exports often use ';'); default ','.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delim = dialect.delimiter
    except csv.Error:
        delim = ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    all_rows = [r for r in reader]
    if not all_rows:
        raise ImportParseError("empty CSV")
    header = [str(h).strip() for h in all_rows[0]]
    data = all_rows[1:]
    return header, data


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportParseError(f"xlsx support unavailable: {exc}") from exc
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportParseError(f"unreadable xlsx: {exc}") from exc
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportParseError("empty xlsx")
    header = [str(h).strip() if h is not None else "" for h in header_row]
    data = [list(r) for r in rows_iter]
    wb.close()
    return header, data


def _cell(v: Any) -> str:
    """Normalize a cell to a trimmed string ('' for None)."""
    if v is None:
        return ""
    return str(v).strip()


def parse_file(
    filename: str,
    content: bytes,
    column_map: dict,
    max_rows: int,
) -> tuple[list[dict], list[int]]:
    """Parse a CSV/xlsx file into Camada-A ingest rows using the mailing's column_map.

    Returns (rows, lines) where lines[i] is the 1-based source line of rows[i] (header
    is line 1, so the first data row is line 2). Fully-empty rows are skipped (not
    counted against the cap). Raises RowCapExceeded / ImportParseError.
    """
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        header, data = _read_csv(content)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        header, data = _read_xlsx(content)
    else:
        raise ImportParseError(f"unsupported file type: {filename!r} (use .csv or .xlsx)")

    if not header:
        raise ImportParseError("missing header row")

    # Cap counts non-empty data rows only.
    non_empty = [(idx, r) for idx, r in enumerate(data)
                 if any(_cell(c) for c in r)]
    if len(non_empty) > max_rows:
        raise RowCapExceeded(len(non_empty), max_rows)

    id_col     = column_map.get("customer_id_column")
    anchors_m  = column_map.get("anchors") or []
    contacts_m = column_map.get("contacts") or {}
    meta_cols  = column_map.get("metadata_columns")

    # Columns consumed by id/anchor/contact mappings — excluded from "rest" metadata.
    used_cols: set[str] = set()
    if id_col:
        used_cols.add(id_col)
    for a in anchors_m:
        if a.get("column"):
            used_cols.add(a["column"])
    for col in contacts_m.values():
        used_cols.add(col)

    col_index = {h: i for i, h in enumerate(header)}

    def _get(row: list[Any], col: str | None) -> str:
        if not col or col not in col_index:
            return ""
        i = col_index[col]
        return _cell(row[i]) if i < len(row) else ""

    rows: list[dict] = []
    lines: list[int] = []
    for data_idx, row in non_empty:
        # customer_id (native)
        customer_id = _get(row, id_col) or None

        # anchors → resolve
        anchors: list[dict] = []
        for a in anchors_m:
            val = _get(row, a.get("column"))
            if val:
                anchors.append({"kind": a.get("kind"), "value": val})

        # contacts (channel → column)
        contacts: dict[str, str] = {}
        for channel, col in contacts_m.items():
            val = _get(row, col)
            if val:
                contacts[channel] = val

        # metadata: listed columns, else all remaining (non-consumed) columns.
        metadata: dict[str, Any] = {}
        target_cols = meta_cols if meta_cols else [h for h in header if h and h not in used_cols]
        for col in target_cols:
            val = _get(row, col)
            if val:
                metadata[col] = val

        rows.append({
            "customer_id": customer_id,
            "anchors":     anchors,
            "contacts":    contacts,
            "metadata":    metadata,
        })
        lines.append(data_idx + 2)   # +1 (0-based→1-based) +1 (header is line 1)

    return rows, lines
