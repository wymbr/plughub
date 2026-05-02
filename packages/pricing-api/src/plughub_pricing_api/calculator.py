"""
calculator.py
PricingCalculator: reads installation resources + Config API price table
→ produces a structured Invoice for a billing cycle.

Pricing model:
  Base pools    — charged for the full cycle regardless of usage.
  Reserve pools — charged per calendar day active (full-day billing).
                  Activated/deactivated manually by the operator.

Invoice structure:
  invoice
    cycle_start, cycle_end, billing_days
    currency
    base_items[]        — one row per (resource_type, installation_id)
    reserve_groups[]    — one group per reserve_pool_id
      pool_id, label, active, days_active, billing_days
      items[]           — one row per resource_type inside the pool
    base_total
    reserve_total
    grand_total
    generated_at
"""
from __future__ import annotations

import calendar
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import asyncpg
import httpx

from . import db as pricing_db

logger = logging.getLogger("plughub.pricing.calculator")

# ─── Default unit prices (used when Config API is unavailable) ─────────────────

DEFAULT_UNIT_PRICES: dict[str, float] = {
    "ai_agent":          120.00,
    "human_agent":        50.00,
    "whatsapp_number":    15.00,
    "voice_trunk_in":     40.00,
    "voice_trunk_out":    40.00,
    "email_inbox":        25.00,
    "sms_number":         10.00,
    "webchat_instance":   20.00,
}
DEFAULT_RESERVE_MARKUP_PCT: float = 0.0   # 0% surcharge on reserve capacity
DEFAULT_CURRENCY: str = "BRL"

# ─── Data model ───────────────────────────────────────────────────────────────

@dataclass
class InvoiceLineItem:
    resource_type: str
    label:         str
    quantity:      int
    unit_price:    float
    days_active:   int | None   # None for base (full cycle)
    billing_days:  int
    subtotal:      float

    def to_dict(self) -> dict:
        return {
            "resource_type": self.resource_type,
            "label":         self.label,
            "quantity":      self.quantity,
            "unit_price":    self.unit_price,
            "days_active":   self.days_active,
            "billing_days":  self.billing_days,
            "subtotal":      round(self.subtotal, 2),
        }


@dataclass
class ReserveGroup:
    pool_id:      str
    label:        str
    active:       bool
    days_active:  int
    billing_days: int
    items:        list[InvoiceLineItem] = field(default_factory=list)

    @property
    def subtotal(self) -> float:
        return sum(i.subtotal for i in self.items)

    def to_dict(self) -> dict:
        return {
            "pool_id":      self.pool_id,
            "label":        self.label,
            "active":       self.active,
            "days_active":  self.days_active,
            "billing_days": self.billing_days,
            "items":        [i.to_dict() for i in self.items],
            "subtotal":     round(self.subtotal, 2),
        }


@dataclass
class Invoice:
    tenant_id:      str
    installation_id: str
    cycle_start:    date
    cycle_end:      date
    billing_days:   int
    currency:       str
    base_items:     list[InvoiceLineItem] = field(default_factory=list)
    reserve_groups: list[ReserveGroup]   = field(default_factory=list)
    generated_at:   datetime             = field(default_factory=datetime.utcnow)

    @property
    def base_total(self) -> float:
        return sum(i.subtotal for i in self.base_items)

    @property
    def reserve_total(self) -> float:
        return sum(g.subtotal for g in self.reserve_groups)

    @property
    def grand_total(self) -> float:
        return self.base_total + self.reserve_total

    def to_dict(self) -> dict:
        return {
            "tenant_id":       self.tenant_id,
            "installation_id": self.installation_id,
            "cycle_start":     self.cycle_start.isoformat(),
            "cycle_end":       self.cycle_end.isoformat(),
            "billing_days":    self.billing_days,
            "currency":        self.currency,
            "base_items":      [i.to_dict() for i in self.base_items],
            "reserve_groups":  [g.to_dict() for g in self.reserve_groups],
            "base_total":      round(self.base_total, 2),
            "reserve_total":   round(self.reserve_total, 2),
            "grand_total":     round(self.grand_total, 2),
            "generated_at":    self.generated_at.isoformat(),
        }


# ─── Calculator ───────────────────────────────────────────────────────────────

class PricingCalculator:
    """
    Stateless calculator. Receives a DB pool and a pre-fetched price table.
    Call `calculate()` to produce an Invoice for a billing cycle.
    """

    def __init__(
        self,
        pg_pool: asyncpg.Pool,
        price_table: dict[str, Any],
    ) -> None:
        self._pool        = pg_pool
        self._prices      = price_table
        self._unit_prices = {**DEFAULT_UNIT_PRICES, **price_table.get("unit_prices", {})}
        self._markup_pct  = float(price_table.get("reserve_markup_pct", DEFAULT_RESERVE_MARKUP_PCT))
        self._currency    = str(price_table.get("currency", DEFAULT_CURRENCY))

    def unit_price(self, resource_type: str) -> float:
        return self._unit_prices.get(resource_type, 0.0)

    def reserve_unit_price(self, resource_type: str) -> float:
        base = self.unit_price(resource_type)
        return base * (1 + self._markup_pct / 100)

    async def calculate(
        self,
        tenant_id: str,
        installation_id: str = "default",
        cycle_start: date | None = None,
        cycle_end:   date | None = None,
    ) -> Invoice:
        """Compute invoice for the given billing cycle (defaults to current month)."""
        today = date.today()
        if cycle_start is None:
            cycle_start = today.replace(day=1)
        if cycle_end is None:
            last_day = calendar.monthrange(cycle_start.year, cycle_start.month)[1]
            cycle_end = cycle_start.replace(day=last_day)

        billing_days = (cycle_end - cycle_start).days + 1

        resources = await pricing_db.list_resources(self._pool, tenant_id, installation_id)

        invoice = Invoice(
            tenant_id=tenant_id,
            installation_id=installation_id,
            cycle_start=cycle_start,
            cycle_end=cycle_end,
            billing_days=billing_days,
            currency=self._currency,
        )

        # ── Base resources (always billed for the full cycle) ──────────────────
        for r in resources:
            if r["pool_type"] != "base":
                continue
            qty   = r["quantity"]
            price = self.unit_price(r["resource_type"])
            invoice.base_items.append(InvoiceLineItem(
                resource_type = r["resource_type"],
                label         = r["label"] or r["resource_type"].replace("_", " ").title(),
                quantity      = qty,
                unit_price    = price,
                days_active   = None,
                billing_days  = billing_days,
                subtotal      = qty * price,
            ))

        # ── Reserve pools (billed per day active in the cycle) ─────────────────
        # Group resources by reserve_pool_id
        reserve_pools: dict[str, list[dict]] = {}
        for r in resources:
            if r["pool_type"] != "reserve":
                continue
            pid = r["reserve_pool_id"] or "reserve"
            reserve_pools.setdefault(pid, []).append(r)

        for pool_id, pool_resources in sorted(reserve_pools.items()):
            days_active = await pricing_db.count_active_days(
                self._pool, tenant_id, pool_id, cycle_start, cycle_end
            )
            # Pool is currently active if any resource has active=True
            currently_active = any(r["active"] for r in pool_resources)
            # Pool label from first resource label or pool_id
            pool_label = pool_resources[0]["label"] or pool_id.replace("_", " ").title()

            group = ReserveGroup(
                pool_id      = pool_id,
                label        = pool_label,
                active       = currently_active,
                days_active  = days_active,
                billing_days = billing_days,
            )

            for r in pool_resources:
                qty        = r["quantity"]
                unit       = self.reserve_unit_price(r["resource_type"])
                daily_rate = unit / billing_days
                subtotal   = daily_rate * qty * days_active
                group.items.append(InvoiceLineItem(
                    resource_type = r["resource_type"],
                    label         = r["resource_type"].replace("_", " ").title(),
                    quantity      = qty,
                    unit_price    = unit,
                    days_active   = days_active,
                    billing_days  = billing_days,
                    subtotal      = subtotal,
                ))

            invoice.reserve_groups.append(group)

        return invoice


# ─── Price table loader ────────────────────────────────────────────────────────

async def load_price_table(config_api_url: str, tenant_id: str) -> dict[str, Any]:
    """
    Fetches the 'pricing' namespace from Config API.
    Falls back to empty dict (defaults will be used) if unavailable.
    """
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            res = await client.get(
                f"{config_api_url}/config/pricing",
                params={"tenant_id": tenant_id},
            )
            if res.status_code == 200:
                data = res.json()
                # Config API returns {key: ConfigEntry} — extract values
                return {k: v["value"] for k, v in data.items() if "value" in v}
    except Exception as exc:
        logger.warning("Could not fetch price table from config-api: %s", exc)
    return {}


# ─── XLSX export ──────────────────────────────────────────────────────────────

def invoice_to_xlsx(invoice: Invoice) -> bytes:
    """Renders invoice as an .xlsx file and returns raw bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ── Colour palette ─────────────────────────────────────────────────────────
    DARK_BG   = "0D1117"
    HEADER_BG = "1E293B"
    BASE_BG   = "0F172A"
    RESERVE_BG = "1A2540"
    TOTAL_BG  = "2563EB"
    WHITE     = "E2E8F0"
    MUTED     = "64748B"
    GREEN     = "22C55E"
    BLUE      = "3B82F6"

    def hdr_font(bold=True, color=WHITE, size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def body_font(color=WHITE, bold=False):
        return Font(name="Calibri", color=color, size=10, bold=bold)

    def fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def money(val: float) -> str:
        return f"{invoice.currency} {val:,.2f}"

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "PlugHub Platform — Invoice"
    ws["A1"].font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    ws["A1"].fill = fill(DARK_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("Tenant",       invoice.tenant_id),
        ("Installation", invoice.installation_id),
        ("Period",       f"{invoice.cycle_start} → {invoice.cycle_end}  ({invoice.billing_days} days)"),
        ("Generated",    invoice.generated_at.strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = body_font(color=MUTED)
        ws[f"A{i}"].fill = fill(DARK_BG)
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = body_font()
        ws[f"B{i}"].fill = fill(DARK_BG)
        for col in "CDEF":
            ws[f"{col}{i}"].fill = fill(DARK_BG)

    row = len(meta) + 2  # spacer

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["Resource", "Type", "Qty", "Unit Price", "Days Active", "Subtotal"]
    col_widths = [28, 18, 8, 14, 14, 16]
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font  = hdr_font()
        cell.fill  = fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Base items ─────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="BASE CAPACITY").font = hdr_font(color=BLUE)
    ws.cell(row=row, column=1).fill = fill(BASE_BG)
    for ci in range(2, 7):
        ws.cell(row=row, column=ci).fill = fill(BASE_BG)
    row += 1

    for item in invoice.base_items:
        data = [
            item.label, item.resource_type,
            item.quantity, money(item.unit_price),
            "full cycle", money(item.subtotal),
        ]
        for ci, val in enumerate(data, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font  = body_font()
            cell.fill  = fill(BASE_BG)
            cell.alignment = Alignment(
                horizontal="center" if ci > 2 else "left",
                vertical="center"
            )
        row += 1

    # Base subtotal row
    ws.cell(row=row, column=5, value="Base Total").font = hdr_font(color=BLUE)
    ws.cell(row=row, column=6, value=money(invoice.base_total)).font = hdr_font(color=BLUE)
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill(BASE_BG)
    row += 2  # spacer

    # ── Reserve groups ─────────────────────────────────────────────────────────
    if invoice.reserve_groups:
        ws.cell(row=row, column=1, value="RESERVE POOLS").font = hdr_font(color=GREEN)
        ws.cell(row=row, column=1).fill = fill(RESERVE_BG)
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).fill = fill(RESERVE_BG)
        row += 1

        for group in invoice.reserve_groups:
            status = "● ACTIVE" if group.active else "○ inactive"
            status_color = GREEN if group.active else MUTED
            ws.cell(row=row, column=1, value=f"  {group.label}").font = hdr_font(bold=True, color=WHITE)
            ws.cell(row=row, column=2, value=status).font = hdr_font(color=status_color, bold=False)
            ws.cell(row=row, column=5, value=f"{group.days_active}/{group.billing_days} days").font = body_font(color=MUTED)
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).fill = fill(RESERVE_BG)
            row += 1

            for item in group.items:
                data = [
                    f"    {item.label}", item.resource_type,
                    item.quantity, money(item.unit_price),
                    f"{item.days_active} days", money(item.subtotal),
                ]
                for ci, val in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font  = body_font()
                    cell.fill  = fill(RESERVE_BG)
                    cell.alignment = Alignment(
                        horizontal="center" if ci > 2 else "left",
                        vertical="center",
                    )
                row += 1

            ws.cell(row=row, column=5, value=f"Pool Total").font = body_font(color=GREEN)
            ws.cell(row=row, column=6, value=money(group.subtotal)).font = body_font(color=GREEN, bold=True)
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).fill = fill(RESERVE_BG)
            row += 1

        row += 1  # spacer

    # ── Grand total ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = hdr_font(bold=True, color=WHITE, size=12)
    ws.cell(row=row, column=6, value=money(invoice.grand_total)).font = hdr_font(bold=True, color=WHITE, size=12)
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill(TOTAL_BG)
    ws.row_dimensions[row].height = 22

    # ── Save to bytes ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
